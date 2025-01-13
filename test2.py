import openpyxl
import datetime
from app import get_file_dict
from pathlib import Path

N_IMG_COL = 3 # 横に並ぶ画像の数
COL_SPACE = 2 # 横にならぶ画像の間隔
ROW_SPACE_SMALL = 2 # 縦に並ぶ同じディレクトリの画像の間隔
ROW_SPACE_BIG = 4 # ディレクトリが切り替わるときの縦の間隔

class ExcelImg:
    """Excelに画像を貼り付ける"""
    def __init__(self, dir_data):
        self.dir_data = dir_data
        self.current = None

def main():
    data = get_file_dict('test.txt')[0]
    dirpath = Path(data.dirpath)
    s = dirpath.name
    print(s)
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws['A1'].value = s
    print(ws['A1'].value)
    book_name = f'{datetime.datetime.now():%Y%m%d%H%M%S}.xlsx'
    wb.save(book_name)
    wb.close()

if __name__ == '__main__':
    main()