from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU as p2e, cm_to_EMU as c2e, pixels_to_points
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker

import math, os, glob, datetime, re
from pathlib import Path

from app import get_file_dict

DPI = 96
IMG_HEIGHT = 400
ROW_HEIGHT_POINT = 13
COL_WIDTH_PIXEL = 67.2 #マジックナンバー
N_COL = 3 #横に並べる画像の数
COL_GAP = 1 #横に並べる画像の間で何列空けるか
ROW_GAP_SMALL = 2 #同じディレクトリ内で改行するとき何行空けるか
ROW_GAP_BIG = 3 #ディレクトリが切り替わるとき何行空けるか

def get_img(img):
    """画像のパスを渡すとリサイズして返す"""
    img = Image(img)
    img.height, img.width = IMG_HEIGHT, img.width * IMG_HEIGHT / img.height
    return img

def get_img_size_in_cell(img):
    """画像を渡すとエクセルで占める行数、列数を返す（切り上げ）"""
    height_per_row = math.ceil(pixels_to_points(img.height, dpi=DPI) / ROW_HEIGHT_POINT)
    width_per_col = math.ceil(img.width / COL_WIDTH_PIXEL)
    return height_per_row, width_per_col


def paste(ws, img, row, col, colof=0, rowof=0):
    """画像をWSに貼り付ける。左上を何行何列にするか指定"""
    cellh = lambda x: c2e((x * 49.77)/99)
    cellw = lambda x: c2e((x * (18.65-1.71))/10)
    coloffset = cellw(colof)
    rowoffset = cellh(rowof)

    h, w = img.height, img.width
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    marker = AnchorMarker(col=col, colOff=coloffset, row=row, rowOff=rowoffset)
    img.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(img) 

def input_cell(ws, value, row, col):
    """WSに文字列を書き込む"""
    ws.cell(row=row, column=col).value = value

def main(dir_tree):
    wb = Workbook()
    ws = wb.active
    for dir_data in dir_tree:
        dirname = dir_data.dirpath.name
        imgs = dir_data.imgs

class ExcelPaster:
    
    def __init__(self):
        self.row = 1
        self.col = 1
        self.dir_idx = 0
        self.wb = Workbook()
        self.ws = self.wb.active
    
    def write_all_dir(self, dir_tree):
        for dir_data in dir_tree:
            self.write_one_dir(dir_data)
            self.row += ROW_GAP_BIG
            
    
    def write_one_dir(self, dir_data):
        """一つのディレクトリのデータを貼り付ける"""
        dir_name = dir_data.dirpath.name
        img_paths = dir_data.imgs
        
        self.input_cell(dir_name, self.row, 1)
        self.row += ROW_GAP_SMALL

        for i, img in enumerate(img_paths):
            img = self.get_img(img)
            paste(img)
            height_per_row, width_per_col = get_img_size_in_cell(img)
            if i % N_COL == 2:
                self.row += height_per_row + ROW_GAP_SMALL
                self.col = 1
            else:
                self.col += width_per_col + COL_GAP
        self.col = 1

    def write_one_dir(self, imgs):
        """画像ファイルパスのリストを渡すと横一列に貼り付ける"""

            
    def strip_extension(self, s):
        """ファイル名から拡張子を消す"""
        pattern = r'(.*)\.[^.]+'
        return re.match(pattern, s).groups()[0]

    
    def paste_image(self, colof=0, rowof=0 ):
        """画像をWSに貼り付ける。左上を何行何列にするか指定
        """

        cellh = lambda x: c2e((x * 49.77)/99)
        cellw = lambda x: c2e((x * (18.65-1.71))/10)
        coloffset = cellw(colof)
        rowoffset = cellh(rowof)

        img = self.get_img(img)
        h, w = img.height, img.width
        size = XDRPositiveSize2D(p2e(w), p2e(h))
        marker = AnchorMarker(col=self.col, colOff=coloffset, row=self.row, rowOff=rowoffset)
        img.anchor = OneCellAnchor(_from=marker, ext=size)
        self.ws.add_image(img) 


    def get_img(self, img):
        """画像のパスを渡すとリサイズして返す"""
        img = Image(img)
        img.height, img.width = IMG_HEIGHT, img.width * IMG_HEIGHT / img.height
        return img

    def get_img_size_in_cell(self,img):
        """画像を渡すとエクセルで占める行数、列数を返す（切り上げ）"""
        height_per_row = math.ceil(pixels_to_points(img.height, dpi=DPI) / ROW_HEIGHT_POINT)
        width_per_col = math.ceil(img.width / COL_WIDTH_PIXEL)
        return height_per_row, width_per_col


    def input_cell(self, value):
        """WSに文字列を書き込む"""
        self.ws.cell(row=self.row, column=self.col).value = value

        

        
        

# def main():
#     wb = Workbook()
#     ws = wb.active


#     img = get_img('test.png')
#     height_per_row, width_per_col = get_img_size_in_cell(img)
#     row_temp = 2
#     col_temp =5
#     paste(ws, img, row_temp, col_temp, 0, 0)
#     input_cell(ws, 'A1のはず', 1, 1)
#     input_cell(ws, '左上のはず', row_temp, col_temp)
#     input_cell(ws, '右上のはず', row_temp, col_temp + width_per_col + 1)
#     input_cell(ws, '右下のはず', row_temp + height_per_row + 1, col_temp + width_per_col + 1)
#     input_cell(ws, '左下のはず', row_temp + height_per_row + 1, col_temp)

#     book_name = f'{datetime.datetime.now():%Y%m%d%H%M%S}.xlsx'
#     wb.save(book_name)

if __name__ == '__main__':
    xlsx_files = glob.glob('*.xlsx')
    for file in xlsx_files:
        os.remove(Path.cwd() / Path(file))
    
    paster = ExcelPaster()
    res = paster.strip_extension('hoge.png')
    print(res)
